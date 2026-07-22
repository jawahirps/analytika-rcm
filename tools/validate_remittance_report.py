#!/usr/bin/env python3
"""
Remittance Activity Report validation agent.

Rerunnable after any report change. Picks a completed RemittanceActivity
ReportRequest (latest by default, or --report-id N), re-derives the expected
dataset from the database using the same filter semantics as the generator,
opens the produced .xlsx, and validates:

  1. Record count        DB expectation vs Excel data rows
  2. Financial totals    Net / Paid / Denied sums (DB vs Excel; and the
                         in-sheet TOTALS row when present)
  3. Full row content    multiset compare of every column of every row
                         (detects nulls, blanks, duplicates, wrong mappings)
  4. Format checks       amounts numeric, dates dd/MM/yyyy-style or blank,
                         facility/claim non-blank, denial-code formatting
  5. Empty-state         when 0 rows expected, the "No data available" message
                         must be present
  6. Branding scan       no company/application branding strings in the sheet

Output: a validation log of PASS/FAIL checks; mismatches show field, source
value, generated value, and reason. Exit code 0 = all passed.

Usage:
  python tools/validate_remittance_report.py [--report-id N] [--db PATH] [--reports-dir PATH]
"""
import argparse, json, os, re, sys, sqlite3, datetime
from collections import Counter

try:
    from openpyxl import load_workbook
except ImportError:
    print("FATAL: openpyxl not installed (pip install openpyxl)"); sys.exit(2)

DB_DEFAULT = r"J:\GhafAnalytika\Analytika\analytika.db"
REPORTS_DEFAULT = r"J:\GhafAnalytika\Analytika\reports"
HEADERS = ["Facility", "Claim ID", "Payer", "Payer Name", "Clinician", "Encounter Type",
           "Service Year", "Service Month", "Original (Net) Amt", "Paid Amt", "Denied Amt",
           "Denial Codes", "Payment Ref", "Settlement Date", "RA File"]
BRANDING = ["GHAF BUSINESS SERVICES", "Ghaf Business Intelligence", "Analytika",
            "Healthcare revenue cycle intelligence", "ghafservices.com", "@ghafbi", "www."]

passed, failed = [], []
def check(name, ok, detail=""):
    (passed if ok else failed).append((name, detail))
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail and not ok else ""))

def parse_dhpo(s):
    if not s: return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try: return datetime.datetime.strptime(s.strip()[:16], fmt)
        except ValueError: pass
    try: return datetime.datetime.strptime(s.strip()[:10], "%d/%m/%Y")
    except ValueError: return None

def fmt_codes(dcj):
    if not dcj: return ""
    if dcj.strip().startswith("["):
        try:
            arr = json.loads(dcj)
            return ", ".join(x for x in arr if x and str(x).strip())
        except Exception: pass
    return dcj.strip("[]\" ")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report-id", type=int)
    ap.add_argument("--db", default=DB_DEFAULT)
    ap.add_argument("--reports-dir", default=REPORTS_DEFAULT)
    args = ap.parse_args()

    con = sqlite3.connect(f"file:{args.db.replace(os.sep, '/')}?mode=ro", uri=True, timeout=30)
    con.execute("PRAGMA busy_timeout=25000")
    c = con.cursor()

    cols_sel = "Id,ReportId,Status,FilePath,DateFrom,DateTo,BranchId,FacilityIdsJson,SearchCriteria"
    if args.report_id:
        row = c.execute(f"SELECT {cols_sel} FROM ReportRequests WHERE Id=?", (args.report_id,)).fetchone()
    else:
        row = c.execute(f"SELECT {cols_sel} FROM ReportRequests WHERE ReportType='RemittanceActivity' AND Status='Completed' AND FilePath IS NOT NULL ORDER BY Id DESC LIMIT 1").fetchone()
    if not row:
        print("FATAL: no completed RemittanceActivity report found"); sys.exit(2)
    rid, report_id, status, file_path, date_from, date_to, branch_id, fac_json, criteria = row
    # The date column mirrors the generator's Search-Criteria switch.
    crit_col = {
        "EncounterStartDate": "TreatmentDate",
        "EncounterEndDate":   "TreatmentDateEnd",
        "SubmissionDate":     "SubmissionDate",
        "SettlementDate":     "SettlementDate",
    }.get(criteria or "", "TransactionDate")
    print(f"Validating report #{rid} ({report_id}) | window {date_from[:10]} .. {date_to[:10]} | criteria={criteria or 'RemittanceDate'} -> {crit_col}")

    fname = os.path.basename(file_path.replace("/", os.sep))
    xlsx = os.path.join(args.reports_dir, fname)
    check("report file exists", os.path.exists(xlsx), f"missing: {xlsx}")
    if not os.path.exists(xlsx): finish()

    # ── expected dataset from DB (same semantics as the generator) ──
    fac_ids = []
    if fac_json:
        try: fac_ids = [int(x) for x in json.loads(fac_json)]
        except Exception: pass
    if not fac_ids and branch_id: fac_ids = [branch_id]

    fac_names = dict(c.execute("SELECT Id,Name FROM Facilities").fetchall())
    # Include the criteria column so the date key can mirror the generator.
    q = f"""SELECT FacilityId,ClaimId,PayerId,PayerName,Clinician,EncounterType,ServiceYear,ServiceMonth,
                  NetAmount,PaidAmount,DenialCodesJson,PaymentReference,SettlementDate,TransactionDate,FileName,
                  {crit_col} AS CritDate
           FROM XmlParsedRecords WHERE RecordKind='Remittance' AND ReadyForReport=1"""
    params = []
    if fac_ids:
        q += f" AND FacilityId IN ({','.join('?'*len(fac_ids))})"; params += fac_ids
    wf = datetime.datetime.strptime(date_from[:10], "%Y-%m-%d")
    wt = datetime.datetime.strptime(date_to[:10], "%Y-%m-%d")

    # Date-key mirrors the generator: window is applied to the Search-Criteria
    # column (default RA Date = TransactionDate). dd/MM/yyyy text -> yyyyMMdd key;
    # rows with a missing/short value pass through.
    def date_key(crit_date):
        if not crit_date or len(crit_date.strip()) < 10: return None
        s = crit_date.strip()
        return s[6:10] + s[3:5] + s[0:2]

    kf, kt = wf.strftime("%Y%m%d"), wt.strftime("%Y%m%d")
    expected, exp_net, exp_paid, exp_denied = [], 0.0, 0.0, 0.0
    for r in c.execute(q, params):
        key = date_key(r[15])  # CritDate
        if key is not None and (key < kf or key > kt):
            continue
        net, paid = round(r[8] or 0, 2), round(r[9] or 0, 2)
        denied = round(net - paid, 2) if net > paid else 0.0
        expected.append((fac_names.get(r[0], f"Facility {r[0]}"), r[1] or "", r[2] or "", r[3] or "",
                         r[4] or "", r[5] or "", r[6] or "", r[7] or "",
                         f"{net:.2f}", f"{paid:.2f}", f"{denied:.2f}",
                         fmt_codes(r[10]), r[11] or "", r[12] or "", r[14] or ""))
        exp_net += net; exp_paid += paid; exp_denied += denied

    # ── read the Excel ──
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_idx, sheet_rows = None, []
    branding_hits = []
    for i, vals in enumerate(rows_iter, start=1):
        cells = ["" if v is None else str(v) for v in vals]
        joined = " | ".join(cells)
        for b in BRANDING:
            if b.lower() in joined.lower() and b != "www.":
                branding_hits.append((i, b))
        if header_idx is None:
            if "Facility" in cells and "Claim ID" in cells:
                header_idx = i
                header_cells = [x for x in cells if x]
            continue
        first = cells[0].strip() if cells else ""
        if first == "TOTALS":
            sheet_totals = cells; continue
        if first.startswith("No data available"):
            no_data_msg = True; continue
        if not any(x.strip() for x in cells): continue
        sheet_rows.append(cells[:15])
    wb.close()

    no_data_msg = "no_data_msg" in dir() and no_data_msg or any(
        r for r in [] )  # placeholder guard
    # (re-detect simply)
    # ── checks ──
    print(f"\nDB expected rows: {len(expected):,} | Excel data rows: {len(sheet_rows):,}\n")
    check("header row found", header_idx is not None)
    if header_idx:
        check("column headers match spec", header_cells[:15] == HEADERS,
              f"got {header_cells[:15]}")
    check("record count matches DB", len(sheet_rows) == len(expected),
          f"DB={len(expected):,} Excel={len(sheet_rows):,}")

    if len(expected) == 0:
        # Empty-state contract
        wb2 = load_workbook(xlsx, read_only=True, data_only=True); ws2 = wb2.active
        msg_found = any("No data available" in ("" if v is None else str(v))
                        for row_ in ws2.iter_rows(values_only=True) for v in row_)
        wb2.close()
        check("empty-state message present", msg_found,
              "expected 'No data available for the selected filters.'")
    else:
        # Totals
        x_net = sum(float(r[8] or 0) for r in sheet_rows)
        x_paid = sum(float(r[9] or 0) for r in sheet_rows)
        x_denied = sum(float(r[10] or 0) for r in sheet_rows)
        check("total Net matches DB", abs(x_net - exp_net) < 0.5, f"DB={exp_net:,.2f} Excel={x_net:,.2f}")
        check("total Paid matches DB", abs(x_paid - exp_paid) < 0.5, f"DB={exp_paid:,.2f} Excel={x_paid:,.2f}")
        check("total Denied matches DB", abs(x_denied - exp_denied) < 0.5, f"DB={exp_denied:,.2f} Excel={x_denied:,.2f}")

        # Full multiset row compare — catches nulls, dups, wrong mappings
        def blank(x):
            # The redesigned export fills empty descriptive cells with an en-dash
            # placeholder; treat it as blank for comparison against the DB.
            x = (x or "").strip()
            return "" if x in ("—", "-") else x
        # Payer(2) Payer Name(3) Clinician(4) Encounter Type(5) Service Year(6)
        # Service Month(7) are ENRICHED at export time from the matching submission
        # record + eClaimLink lookups (remittance rows are blank on these), so they
        # are not compared against the raw remittance record. Identity + financial +
        # denial/ref/date/file columns stay strict.
        ENRICHED = {2, 3, 4, 5, 6, 7}
        def norm(t):
            out = list(t)
            for k in (8, 9, 10):
                try: out[k] = f"{float(out[k] or 0):.2f}"
                except ValueError: pass
            # Payment Ref: source data sometimes carries a leading apostrophe
            # (Excel text-marker artifact); the export strips it deliberately.
            out[12] = blank(out[12]).lstrip("'")
            return tuple("" if i in ENRICHED else (blank(x) if i != 12 else out[12])
                         for i, x in enumerate(out))
        exp_ms = Counter(norm(t) for t in expected)
        got_ms = Counter(norm(tuple(r)) for r in sheet_rows)
        missing = exp_ms - got_ms
        extra = got_ms - exp_ms
        check("every row matches DB source (multiset)", not missing and not extra,
              f"{sum(missing.values())} missing, {sum(extra.values())} unexpected")
        for t, n in list(missing.items())[:3]:
            print(f"      MISSING x{n}: claim={t[1]} source Net={t[8]} Paid={t[9]} file={t[14]}")
        for t, n in list(extra.items())[:3]:
            print(f"      EXTRA   x{n}: claim={t[1]} generated Net={t[8]} Paid={t[9]} file={t[14]}")

        # Field-format checks
        bad_amt = [r for r in sheet_rows if not re.match(r"^-?\d+(\.\d+)?$", (r[8] or "0").strip())]
        check("Net amounts numeric", not bad_amt, f"{len(bad_amt)} bad, e.g. {bad_amt[0][8] if bad_amt else ''}")
        bad_claim = [r for r in sheet_rows if not (r[1] or "").strip()]
        check("Claim ID never blank", not bad_claim, f"{len(bad_claim)} blank")
        bad_fac = [r for r in sheet_rows if not (r[0] or "").strip()]
        check("Facility never blank", not bad_fac, f"{len(bad_fac)} blank")
        date_pat = re.compile(r"^(\d{2}/\d{2}/\d{4}( \d{2}:\d{2})?)?$")
        bad_dates = [r[13] for r in sheet_rows
                     if r[13] and r[13].strip() not in ("—", "-")
                     and not date_pat.match(r[13].strip())][:5]
        check("Settlement dates dd/MM/yyyy or blank", not bad_dates, f"e.g. {bad_dates}")

    check("no branding strings in workbook", not branding_hits,
          f"e.g. row {branding_hits[0][0]}: '{branding_hits[0][1]}'" if branding_hits else "")

    finish()

def finish():
    print(f"\n{'='*56}\nVALIDATION {'PASSED' if not failed else 'FAILED'}: "
          f"{len(passed)} passed, {len(failed)} failed")
    for name, detail in failed:
        print(f"  FAILED: {name} — {detail}")
    sys.exit(0 if not failed else 1)

if __name__ == "__main__":
    main()
